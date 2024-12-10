import openpyxl
import json


def write_base_doors_to_excel(book):
    sheet_1 = book.create_sheet('Base_series')

    name_path = 'attributes bunker doors for json'
    file_base_door = 'Base_door'

    with open(f'{name_path}/{file_base_door}.json', 'r', encoding='utf8') as file:
        recieve_data = json.load(file)

    headers = ['name', 'price', 'Толщина стали, полотна', 'Конструкция короба', 'Утеплитель двери', 'Замок нижний',
        'Замок верхний', 'Наличники (снаружи)', 'Внутренняя панель', 'Цвет металла', 'Фурнитура', 'link'
    ]

    colunm_headers = 1

    for title_column in headers:
        sheet_1.cell(row=1, column=colunm_headers).value = title_column
        colunm_headers += 1

    row = 2
    column = 1

    for name_series, attributes_doors in recieve_data.items():
        for attribut_door in attributes_doors:
            column = 1
            for attribut, description in attribut_door.items():
                sheet_1.cell(row=row, column=column).value = description
                column += 1
            row += 1

    book.save('Bunker_doors.xlsx')
    book.close()


def write_hit_doors_to_excel(book):
    sheet_2 = book.create_sheet('Hit_series')
    name_path = 'attributes bunker doors for json'
    file_hit_door = 'Hit_door'

    with open(f'{name_path}/{file_hit_door}.json', 'r', encoding='utf8') as file:
        recieve_data = json.load(file)

    headers = ['name', 'price', 'Толщина стали, полотна', 'Контуры уплотнения', 'Конструкция короба', 'Утеплитель двери',
        'Замок нижний', 'Замок верхни', 'Внешняя панель', 'Наличники (снаружи)', 'Внутренняя панель','Цвет металла',
        'Петли', 'Фурнитура', 'link'
    ]

    colunm_headers = 1

    for title_column in headers:
        sheet_2.cell(row=1, column=colunm_headers).value = title_column
        colunm_headers += 1

    row = 2
    column = 1

    for name_series, attributes_doors in recieve_data.items():
        for attribut_door in attributes_doors:
            column = 1
            for attribut, description in attribut_door.items():
                sheet_2.cell(row=row, column=column).value = description
                column += 1
            row += 1

    book.save('Bunker_doors.xlsx')
    book.close()


def write_prime_doors_to_excel(book):
    sheet_3 = book.create_sheet('Prime_series')
    name_path = 'attributes bunker doors for json'
    file_prime_door = 'Prime_door'

    with open(f'{name_path}/{file_prime_door}.json', 'r', encoding='utf8') as file:
        recieve_data = json.load(file)

    headers = ['name', 'price', 'Толщина стали, полотна', 'Контуры уплотнения', 'Конструкция короба', 'Утеплитель двери',
        'Замок нижний', 'Замок верхний', 'Внешняя панель', 'Наличники (снаружи)', 'Внутренняя панель', 'Цвет металла',
        'Петли', 'Фурнитура', 'Дополнительно', 'Особенности', 'link'
    ]

    colunm_headers = 1

    for title_column in headers:
        sheet_3.cell(row=1, column=colunm_headers).value = title_column
        colunm_headers += 1

    row = 2
    column = 1

    for name_series, attributes_doors in recieve_data.items():
        for attribut_door in attributes_doors:
            column = 1
            for attribut, description in attribut_door.items():
                sheet_3.cell(row=row, column=column).value = description
                column += 1
            row += 1

    book.save('Bunker_doors.xlsx')
    book.close()


def write_termo_doors_to_excel(book):
    sheet_4 = book.create_sheet('Termo_series')
    name_path = 'attributes bunker doors for json'
    file_termo_door = 'Termo_door'

    with open(f'{name_path}/{file_termo_door}.json', 'r', encoding='utf8') as file:
        recieve_data = json.load(file)

    headers = ['name', 'price', 'Терморазрыв', 'Толщина стали, полотна', 'Контуры уплотнения', 'Конструкция короба',
        'Утеплитель двери', 'Замок нижний', 'Замок верхний', 'Наличники (снаружи)', 'Внутренняя панель', 'Цвет металла',
        'Петли', 'Фурнитура', 'Особенности', 'link'
    ]
    colunm_headers = 1

    for title_column in headers:
        sheet_4.cell(row=1, column=colunm_headers).value = title_column
        colunm_headers += 1


    for name_series, attributes_doors in recieve_data.items():
        for attribut_door in attributes_doors[0:1]:
            sheet_4.cell(row=2, column=1).value = attribut_door.get('name')
            sheet_4.cell(row=2, column=2).value = attribut_door.get('price')
            sheet_4.cell(row=2, column=4).value = attribut_door.get('Толщина стали, полотна')
            sheet_4.cell(row=2, column=16).value = attribut_door.get('link')

    row = 3
    column = 1

    for name_series, attributes_doors in recieve_data.items():
        for attribut_door in attributes_doors[1:]:
            column = 1
            for attribut, description in attribut_door.items():
                sheet_4.cell(row=row, column=column).value = description
                column += 1
            row += 1

    book.save('Bunker_doors.xlsx')
    book.close()


def write_all_doors_to_excel():
    book = openpyxl.Workbook()
    book.remove(book.active)

    write_base_doors_to_excel(book)
    write_hit_doors_to_excel(book)
    write_prime_doors_to_excel(book)
    write_termo_doors_to_excel(book)


